import requests
from requests.auth import HTTPBasicAuth
import sys
from datetime import datetime
import re
import os
import chinese_calendar
from config import username, password
from setting import groupJiraRegexp, localJiraRegexp, local_jira_url, manual_workdays
import pandas as pd
from openpyxl.styles import PatternFill
import openpyxl
from tabulate import tabulate
from shared.jira import copy_group_to_local


def get_month_range(year, month):
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1) - pd.Timedelta(days=1)
    else:
        end = datetime(year, month + 1, 1) - pd.Timedelta(days=1)
    return start, end


def getExistingWorklogs(year, month):
    month_start, month_end = get_month_range(year, month)
    start_str = month_start.strftime("%Y-%m-%d")
    end_str = month_end.strftime("%Y-%m-%d")
    jql = f'project=11859 and assignee=currentUser() and worklogDate >= "{start_str}" and worklogDate <= "{end_str}"'
    logs_url = f"{local_jira_url}/rest/api/2/search/jql?fields=worklog,assignee,summary&jql={jql}"
    print("🚀 Fetching existing workLogs:", logs_url)
    response = requests.get(
        logs_url,
        auth=HTTPBasicAuth(username, password),
        headers={"Accept": "application/json"},
        timeout=30,
    )
    response.raise_for_status()  # Raise an HTTPError for bad responses (4xx and 5xx)
    issues = response.json().get("issues", [])
    print(f"✅ Fetched {len(issues)} worklogs from local Jira.")
    return issues


def outputExcel(year, month, excel_path):
    try:
        if os.path.exists(excel_path):
            print(f"⚠️ 文件已存在: {excel_path}")
            confirm = input("是否覆盖？(y/n): ").strip().lower()
            if confirm != "y":
                print("❌ 已取消操作。")
                return
        month_start, month_end = get_month_range(year, month)
        issues = getExistingWorklogs(year, month)
        data = []
        for issue in issues:
            key = issue.get("key", "")
            summary = issue.get("fields", {}).get("summary", "")
            assignee = (
                issue.get("fields", {}).get("assignee", {}).get("displayName", "")
            )
            worklogs = issue.get("fields", {}).get("worklog", {}).get("worklogs", [])
            for log in worklogs:
                author = log.get("author", {}).get("displayName", "")
                time_spent = log.get("timeSpentSeconds", 0)
                started = log.get("started", "")
                try:
                    started_dt = datetime.strptime(started[:10], "%Y-%m-%d")
                except Exception:
                    continue
                if month_start <= started_dt <= month_end:
                    data.append(
                        {
                            "Date": started_dt.strftime("%Y-%m-%d"),
                            "Worklog Author": author,
                            "Time Spent (hours)": time_spent / 3600,
                            "Issue Key": key,
                            "Summary": summary,
                            "Assignee": assignee,
                        }
                    )

        df = pd.DataFrame(data)
        weekday_map = {
            0: "周一",
            1: "周二",
            2: "周三",
            3: "周四",
            4: "周五",
            5: "周六",
            6: "周日",
        }

        # 生成本月所有日期
        all_dates = pd.date_range(month_start, month_end)
        date_weekday = [
            f"{d.strftime('%Y-%m-%d')} {weekday_map[d.weekday()]}" for d in all_dates
        ]
        calendar_df = pd.DataFrame({"Date": date_weekday})

        # 记录哪些行是节假日
        holiday_flags = [
            (
                not d.strftime("%Y-%m-%d") in manual_workdays
                and (
                    chinese_calendar.is_holiday(d)
                    or (d.weekday() >= 5)  # and not chinese_calendar.is_workday(d)
                )
            )
            for d in all_dates
        ]

        # 构造 Issue Time Spent 列
        if not df.empty and "Date" in df.columns:
            # 按日期和 issue 聚合
            df["Time Spent (hours)"] = df["Time Spent (hours)"].round(2)
            grouped = (
                df.groupby(["Date", "Issue Key"])["Time Spent (hours)"]
                .sum()
                .reset_index()
            )

            # 拼接成 issueKey:hours 格式（去掉.0）
            def format_hours(x):
                hours = x["Time Spent (hours)"]
                if hours == int(hours):
                    return f"{x['Issue Key']}:{int(hours)}"
                else:
                    return f"{x['Issue Key']}:{hours}"

            grouped["IssueTime"] = grouped.apply(format_hours, axis=1)
            # 按日期合并
            issue_time = (
                grouped.groupby("Date")["IssueTime"]
                .apply(lambda x: ", ".join(x))
                .reset_index()
            )
            # 合并到日历（用中文星期）
            issue_time["Date"] = (
                issue_time["Date"]
                + " "
                + pd.to_datetime(issue_time["Date"]).dt.weekday.map(weekday_map)
            )
            calendar_df = calendar_df.merge(issue_time, on="Date", how="left")
            calendar_df = calendar_df.fillna({"IssueTime": ""})
        else:
            calendar_df["IssueTime"] = ""

        # 只保留两列，并设置 IssueTime 列头
        calendar_df = calendar_df[["Date", "IssueTime"]]
        calendar_df.rename(
            columns={"IssueTime": "Issue Log Time (sample XXXX:8) 单位：hour"},
            inplace=True,
        )

        # 统计总小时数
        total_hours = df["Time Spent (hours)"].sum() if not df.empty else 0
        # 统计预期工时（工作日天数 × 8）
        workday_count = sum([not h for h in holiday_flags])
        expected_hours = workday_count * 8
        print(
            f"💡 本月工作日天数: {workday_count}, 总天数: {len(all_dates)}, 预期工时: {expected_hours} 小时"
        )

        print("💡 本月放假日期：")
        for idx, is_holiday in enumerate(holiday_flags):
            if is_holiday:
                d = all_dates[idx]
                print(f"   📅 {d.strftime('%Y-%m-%d')} {weekday_map[d.weekday()]}")

        # 插入 Summary 列
        calendar_df.insert(2, "Summary", "")

        # 第一行：本月总计
        calendar_df.at[0, "Summary"] = f"本月总计: {total_hours:.2f} 小时"

        # 第二行：预期工时（去掉小数点）
        if expected_hours == int(expected_hours):
            calendar_df.at[1, "Summary"] = f"预期工时: {int(expected_hours)} 小时"
        else:
            calendar_df.at[1, "Summary"] = f"预期工时: {expected_hours:.2f} 小时"

        calendar_df.to_excel(excel_path, index=False)
        print(f"📁 Excel 日历已生成：{excel_path}")

        # 用 openpyxl 设置节假日行灰色
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        ws.column_dimensions["A"].width = 20  # 日期列宽
        ws.column_dimensions["B"].width = 50  # Issue Log Time列宽
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12  # HourSum列宽
        # ws.column_dimensions['D'].hidden = True  # 隐藏HourSum列

        # 统计数据区间
        data_start_row = 3
        data_end_row = ws.max_row

        # D列：HourSum，每行只取第一个工时
        ws["D1"] = "HourSum"
        for row in range(2, data_end_row + 1):
            ws[f"D{row}"] = (
                f'=IFERROR(IFERROR(VALUE(MID(SUBSTITUTE(B{row}," ",""),FIND(":",SUBSTITUTE(B{row}," ",""))+1,IFERROR(FIND(",",SUBSTITUTE(B{row}," ","")&",")-FIND(":",SUBSTITUTE(B{row}," ",""))-1,99))),0)'
                f'+IFERROR(IF(FIND(",",SUBSTITUTE(B{row}," ",""),FIND(",",SUBSTITUTE(B{row}," ",""))+1)>0,VALUE(MID(SUBSTITUTE(B{row}," ",""),FIND(":",SUBSTITUTE(B{row}," ",""),FIND(",",SUBSTITUTE(B{row}," ",""),FIND(",",SUBSTITUTE(B{row}," ",""))+1))+1,99)),0),0)'
                f'+IFERROR(IF(FIND(",",SUBSTITUTE(B{row}," ",""))>0,VALUE(MID(SUBSTITUTE(B{row}," ",""),FIND(":",SUBSTITUTE(B{row}," ",""),FIND(",",SUBSTITUTE(B{row}," ","")))+1,IFERROR(FIND(",",SUBSTITUTE(B{row}," ","")&",",FIND(",",SUBSTITUTE(B{row}," ",""))+1)-FIND(":",SUBSTITUTE(B{row}," ",""),FIND(",",SUBSTITUTE(B{row}," ","")))-1,99))),0),0)'
                f",0)"
            )

        # Summary 统计直接放在 C 列
        ws["C1"] = "Summary"
        ws["C2"] = f'="当前工时: " & SUM(D2:D{data_end_row}) & " 小时"'

        if expected_hours == int(expected_hours):
            ws["C3"] = f"预期工时: {int(expected_hours)} 小时"
        else:
            ws["C3"] = f"预期工时: {expected_hours:.2f} 小时"

        # 只在第2、3行有公式，其他行为空
        for row in range(4, data_end_row + 1):
            ws[f"C{row}"] = ""

        # 灰色节假日行
        for idx, is_holiday in enumerate(holiday_flags, start=2):
            if is_holiday:
                for cell in ws[idx]:
                    cell.fill = PatternFill("solid", fgColor="DDDDDD")

        wb.save(excel_path)
        print("🎨 节假日已用灰色标记")
    except requests.exceptions.RequestException as e:
        print(f"❌❌ An error occurred: {e}")


def updateLogwork(worklog_id, issue_key, log_date, hours):
    url = f"{local_jira_url}/rest/api/2/issue/{issue_key}/worklog/{worklog_id}"
    payload = {
        "started": f"{log_date}T09:00:00.000+0000",  # 假设工时从上午9点开始
        "timeSpentSeconds": int(hours * 3600),
        "comment": f"自动记录工时: {hours} 小时",
    }
    response = requests.put(
        url,
        auth=HTTPBasicAuth(username, password),
        headers={"Content-Type": "application/json"},
        json=payload,
    )
    if response.status_code == 204:
        print(f"✅ 工时已更新: {issue_key} | {log_date} | {hours} 小时")
    else:
        print(f"❌ 工时更新失败: {response.status_code}, {response.text}")


def deleteLogWork(worklog_id, issue_key):
    url = f"{local_jira_url}/rest/api/2/issue/{issue_key}/worklog/{worklog_id}"
    response = requests.delete(
        url,
        auth=HTTPBasicAuth(username, password),
        headers={"Content-Type": "application/json"},
    )
    if response.status_code == 204:
        print(f"✅ 工时已删除: {issue_key} | worklogId: {worklog_id}")
    else:
        print(f"❌ 工时删除失败: {response.status_code}, {response.text}")


def addLogWork(issue_key, log_date, hours):
    url = f"{local_jira_url}/rest/api/2/issue/{issue_key}/worklog"
    payload = {
        "started": f"{log_date}T09:00:00.000+0000",  # 假设工时从上午9点开始
        "timeSpentSeconds": int(hours * 3600),
        "comment": f"自动记录工时: {hours} 小时",
    }
    response = requests.post(
        url,
        auth=HTTPBasicAuth(username, password),
        headers={"Content-Type": "application/json"},
        json=payload,
    )
    if response.status_code == 201:
        print(f"✅ 工时已记录: {issue_key} | {log_date} | {hours} 小时")
    else:
        print(f"❌ 工时记录失败: {response.status_code}, {response.text}")


## 记录工时
def logwork(year, month, result):
    # existingWorklogs = getExistingWorklogs(year, month)
    localRes = []
    for item in result:
        date = item["date"]
        worklog = item["worklog"]
        isGroupJira = item["isGroupJira"]
        localKey = item["issue"]
        if isGroupJira:
            # 处理GroupJira工时记录
            print(f"🏁🏁 开始处理 GroupJira to Local: {localKey} ...")
            localKey = copy_group_to_local(
                item["issue"],
                updateStatus=True,
                uploadAttachment=False,
                uploadComment=False,
                skipUpdate=True,
            )
        localRes.append(
            {
                "date": date,
                "issue": localKey,
                "worklog": worklog,
                "isGroupJira": isGroupJira,
            }
        )
    if not localRes:
        print("❌ 没有有效的工时记录，退出。")
        sys.exit(1)

    existingWorklogs = getExistingWorklogs(year, month)
    # localRes [{'date': '2025-10-01 周三', 'issue': 'OMNE-148', 'worklog': '8', 'isGroupJira': True}]
    for workItem in localRes:
        # print("📝 工时记录:", workItem)
        # 判断existingWorklogs是否已经存在当天的关于这个jira的记录，如果相同就跳过，如果有不同就update，没有就新增，其他的就删除
        log_date = workItem["date"][:10]
        issue_key = workItem["issue"]
        hours = float(workItem["worklog"])
        isGroupJira = workItem["isGroupJira"]
        if hours <= 0:
            print(f"⚠️ 工时必须大于0，跳过: {workItem}")
            continue
        if not re.match(localJiraRegexp, issue_key):
            print(f"❌ 非法的本地Jira Key，跳过: {issue_key}")
            continue
        # 查找是否已有记录
        # 查找当天该issue的所有worklog
        logs_to_update = []
        total_logged_hours = 0
        for issue in existingWorklogs:
            if issue.get("key") != issue_key:
                continue
            worklogs = issue.get("fields", {}).get("worklog", {}).get("worklogs", [])
            for log in worklogs:
                started = log.get("started", "")[:10]
                if started == log_date:
                    logs_to_update.append(log)
                    total_logged_hours += log.get("timeSpentSeconds", 0) / 3600
        if abs(total_logged_hours - hours) < 0.01 and logs_to_update:
            print(f"⏩ 已有相同工时记录，跳过: {issue_key} {log_date}")
        elif len(logs_to_update) == 1:
            print(
                f"🔄 单条工时不同，直接更新: {issue_key} {log_date} {total_logged_hours}→{hours}"
            )
            updateLogwork(logs_to_update[0].get("id"), issue_key, log_date, hours)
        elif logs_to_update:
            print(
                f"🔄 多条工时，合并更新: {issue_key} {log_date} {total_logged_hours}→{hours}"
            )
            for log in logs_to_update:
                print(f"⏩ 删除: {issue_key} {log_date} worklogId={log.get('id')}")
                deleteLogWork(log.get("id"), issue_key)
            print(f"➕ 新增工时记录: {issue_key} {log_date} {hours}")
            addLogWork(issue_key, log_date, hours)
        else:
            print(f"➕ 新增工时记录: {issue_key} {log_date} {hours}")
            addLogWork(issue_key, log_date, hours)

    # 删除多余的工时（本地没有但Jira有的）
    for issue in existingWorklogs:
        issue_key = issue.get("key")
        worklogs = issue.get("fields", {}).get("worklog", {}).get("worklogs", [])
        for log in worklogs:
            started = log.get("started", "")[:10]
            author = log.get("author", {}).get("name", "")
            # 如果当天该issue不在localRes，则删除
            if not any(
                w["issue"] == issue_key and w["date"][:10] == started for w in localRes
            ):
                print(
                    f"🗑️  删除多余工时: {issue_key} {started} worklogId={log.get('id')}"
                )
                deleteLogWork(log.get("id"), issue_key)


def inputExcel(year, month, inputPath):
    if not os.path.exists(inputPath):
        print(f"❌ 未找到文件: {inputPath}\n请先使用 --output 参数生成Excel文件")
        sys.exit(1)
    else:
        print(f"✅ 找到本地Excel文件: {inputPath}")
        df = pd.read_excel(inputPath)
        print("📖 Excel内容预览：")
        if (
            "Date" in df.columns
            and "Issue Log Time (sample XXXX:8) 单位：hour" in df.columns
        ):
            result = []
            unknown_issues = []
            for idx, row in df.iterrows():
                date = row["Date"]
                issue_times = row["Issue Log Time (sample XXXX:8) 单位：hour"]
                if issue_times:
                    for item in str(issue_times).split(","):
                        item = item.strip()
                        if ":" in item:
                            issue, hours = item.split(":")
                            is_group_jira = bool(re.match(groupJiraRegexp, issue))
                            is_local_jira = bool(re.match(localJiraRegexp, issue))
                            if not is_group_jira and not is_local_jira:
                                unknown_issues.append(issue)
                            result.append(
                                {
                                    "date": date,
                                    "issue": issue,
                                    "worklog": hours,
                                    "isGroupJira": is_group_jira,
                                }
                            )
                            # print(
                            #     f"日期: {date} | {issue} {'(GroupJIRA)' if is_group_jira else ''} | 工时: {hours}"
                            # )
            if unknown_issues:
                print(
                    f"\n❌ 以下Issue不在groupJira或localJira范围，请检查：\n{', '.join(unknown_issues)}"
                )
                sys.exit(1)
            if not result:
                print("❌ 没有有效的工时记录，退出。")
                sys.exit(1)

            if result:
                print(tabulate(result, headers="keys", tablefmt="psql", showindex=True))
            print(f"\n✅ 共解析 {len(result)} 条记录，是否确认？(y/n)")
            confirm = input().strip().lower()
            if confirm == "y":
                print("✔️ 已确认，继续后续操作。")
                logwork(year, month, result)
            else:
                print("❌ 已取消操作。")
                sys.exit(0)
        else:
            print("Excel缺少必要的列，无法解析。")
